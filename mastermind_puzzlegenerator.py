#!/usr/bin/env python3

"""Generates a set of mastermind to solve i.e. an Excel-workbook to print out
with both participants' sheets and the stops.
"""

import argparse
import openpyxl
import openpyxl.styles
import random
from openpyxl.styles import Color, PatternFill, Border, Side, Font, \
    Alignment
from functools import reduce

HEADING_PER_SHEET = "Lagblankett (svår)"
HEADING_PER_EASY_SHEET = "Lagblankett"
HEADING_CORRECT_ANSWERS = "Facit"

INTRO_TEXT_PER_SHEET = """\
Det här är lagblanketten för Radiomastermind på Skogsrå.

Bergatrollet är väldigt olyckligt för någon, troligtvis knytten, har
ändrat koden för att komma in till hans skattgömma. Han har, i största
hemlighet, kontaktat oss för att vi har magiska krafter och apparater
som kan hjälpa. Vi har lyckats lista ut hur man ska få fram koden men
behöver er hjälp för att Bergatrollet ska kunna komma åt sina
rikedomar.

Koden som saknas är en kombination av färger med nummer. Med våra
magiska metoder, kan vi tvinga Knytten att lämna ledtrådar till koder
som vi provar. Vi vet att det fungerar och att någon, troligen
Knytten, lämnar denna information längs Knyttstigen men vi har inte
sett något knytt göra det.

Detta papper innehåller koder som vi provat med. Hur bra varje kod är
framgår av antalet svarta och vita. Svarta anger hur många av kodens
färger som återfinns i den rätta koden på rätt plats. Antalet vita
anger hur många av som återfinns på fel plats. Det gäller att använda
informationen om svarta och vita för att få fram rätt kod.

Det är farligt att samla in dessa ledtrådar. Bara en scout skyddas av
apparatens magi så ni måste skicka ut en modig scout som använder
apparaten för att söka efter ledtrådar samtidigt som ni andra använder
informationen för att så snabbt som möjligt lista ut koden."""

CORRECT_HEADING = """Rätt kod:"""
SOLVED_IN_HEADING = """Löses på"""

STOP_HEADING = """kontroll"""
BLACK_HEADING = """svarta"""
WHITE_HEADING = """vita"""
CLUE_HEADING = """ledtråd"""

HEADING_PER_STOP = "Radiomastermind kontroll nummer"

INTRO_TEXT_PER_STOP = """Detta är en kontroll för Radiomastermind på Skogsrå.

Kontrollen innehåller ledtrådar till att hitta koden till
Bergatrollets skattgömma. Ni kan hjälpa Bergatrollet med detta hos
radioscouterna."""

ROWS_PER_SHEET = 58

INTRO_ALIGNMENT = Alignment(vertical="top",
                            wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center")
STOP_ALIGNMENT = CENTER_ALIGNMENT
CLUE_ALIGNMENT = CENTER_ALIGNMENT
COLOR_ALIGNMENT = CENTER_ALIGNMENT
HEADING_FONT = Font(size=9, bold=True)
SIDE = Side(border_style="thin",
            color='FF000000')

HEADING_SIDE = Side(border_style="double",
                    color='FF000000')
HEADER_BORDER = Border(top=HEADING_SIDE,
                       bottom=HEADING_SIDE,
                       left=HEADING_SIDE,
                       right=HEADING_SIDE)
CELL_BORDER = Border(top=SIDE,
                     bottom=SIDE,
                     left=SIDE,
                     right=SIDE)

parser = argparse.ArgumentParser(
    description="Generate a set of mastermind games.")
parser.add_argument('--sheets', type=int, default=2,
                    help='Number of sheets (counting also the easy ones)')
parser.add_argument('--easy', type=int, default=0,
                    help='Number of the sheets that are easy')
parser.add_argument('--columns', type=int, default=4,
                    help='Number of clues to guess')
parser.add_argument('--colors', type=int, default=8,
                    help='Number of colors to choose from')
parser.add_argument('--stops', type=int, default=15,
                    help='Number of stops to go to on the course')
parser.add_argument('--filename', type=str,
                    help='The file where the result is stored',
                    default="mastermind.xlsx")
parser.add_argument('--debug', '-d', action='store_true',
                    help='Activate trace outputs',
                    default=False)


def random_line(args):
    """Generate a random line."""
    line = []
    for _ in range(args.columns):
        line.append(random.randint(1, args.colors))
    return line


class TooManyClues(Exception):
    pass


class NoCombinationsLeft(Exception):
    pass


class Sheet(object):
    def __init__(self, args, easy=False):
        """Creates a sheet."""
        self.args = args
        self.easy = easy
        self.correct = random_line(self.args)
        self.clue_lines = []
        self.clue_answers = []
        combs = self.combinations(self.clue_lines)
        while combs > 1:
            if len(self.clue_lines) >= self.args.stops:
                raise TooManyClues()
            new_line = random_line(self.args)
            if new_line == self.correct:
                # Too easy
                continue
            if self.easy:
                if self.answer(new_line)[0] == 0:
                    continue
            new_clue_lines = self.clue_lines + [new_line]
            new_combs = self.combinations(new_clue_lines)
            if new_combs < combs:
                self.clue_lines = new_clue_lines
                self.clue_answers.append(self.answer(new_line))
                combs = new_combs
        self.solvable = len(self.clue_lines)
        if self.args.debug:
            print("Verified that the sheet is solvable.",
                  self.solvable, "lines.")
        while len(self.clue_lines) < self.args.stops:
            new_line = random_line(self.args)
            self.clue_lines.append(new_line)
            self.clue_answers.append(self.answer(new_line))

    def answer(self, clue_line, correct=None):
        """Returns a tuple of counts for black and white."""
        if correct is None:
            correct = self.correct
        count_black = 0
        rest_correct = list(correct)
        rest_clue = list(clue_line)
        for i in range(self.args.columns):
            if correct[i] == clue_line[i]:
                count_black += 1
                rest_correct.remove(correct[i])
                rest_clue.remove(clue_line[i])
        count_white = 0
        for c in rest_clue:
            if c in rest_correct:
                count_white += 1
                rest_correct.remove(c)

        return count_black, count_white

    def combinations(self, clue_lines):
        reduced_combinations = [set(range(1, self.args.colors + 1))
                                for _ in range(self.args.columns)]
        for clue_line in clue_lines:
            black, white = self.answer(clue_line)
            if black == 0 and white == 0:
                for i in range(self.args.columns):
                    for j in range(self.args.columns):
                        if clue_line[j] in reduced_combinations[i]:
                            reduced_combinations[i].remove(clue_line[j])
            elif black == 0:
                for i in range(self.args.columns):
                    if clue_line[i] in reduced_combinations[i]:
                        reduced_combinations[i].remove(clue_line[i])
        if self.args.debug:
            red = reduce((lambda x, y: x * y),
                         [len(s) for s in reduced_combinations])
            if red < self.args.columns * self.args.colors:
                print("Reduced combinations:", red)
        combinations = []
        for x in reduced_combinations[0]:
            combinations.append([x])
        for c in range(1, self.args.columns):
            old_combinations = combinations
            combinations = []
            for x in reduced_combinations[c]:
                for comb in old_combinations:
                    combinations.append(comb + [x])
        valid_combinations = []
        for comb in combinations:
            for clue_line in clue_lines:
                if self.answer(clue_line, comb) != self.answer(clue_line):
                    break
            else:
                valid_combinations.append(comb)
        if len(valid_combinations) == 0:
            raise NoCombinationsLeft()
        if self.args.debug:
            print("Combinations:", len(valid_combinations))
        return len(valid_combinations)

    def output(self, ws, sheet_identity, start_row, replacement):
        """Will fill the worksheet from line start_line with the sheet.

        After a common header with explanation, there board is generated
        with each line generated with the clue
        ----------------------------------------------------------------------
        | stop# | color | color | color | ... |      | empty | empty | clue  |
        ----------------------------------------------------------------------

        The amount of color columns depends on the given columns.
        The empty columns are to enter "black" and "white" responses.
        Headers are created.
        """
        stop_column = 1
        black_column = 1 + self.args.columns + 2
        white_column = black_column + 1
        clue_column = white_column + 1

        ws.merge_cells(start_row=start_row, end_row=start_row,
                       start_column=1, end_column=9)
        heading = HEADING_PER_SHEET
        if self.easy:
            heading = HEADING_PER_EASY_SHEET
        ws.cell(row=start_row, column=1).value = heading + " " + sheet_identity
        ws.cell(row=start_row, column=1).alignment = INTRO_ALIGNMENT

        row = start_row + 3
        intro_lines = 22
        ws.merge_cells(start_row=row, end_row=row + intro_lines,
                       start_column=1, end_column=9)
        ws.cell(row=row, column=1).value = INTRO_TEXT_PER_SHEET
        ws.cell(row=row, column=1).alignment = INTRO_ALIGNMENT

        row += intro_lines + 2
        ws.cell(row=row, column=1).value = CORRECT_HEADING
        for column in range(self.args.columns):
            ws.cell(row=row, column=2 + column).border = HEADER_BORDER

        row += 2
        ws.cell(row=row, column=stop_column).value = STOP_HEADING
        ws.cell(row=row, column=stop_column).border = HEADER_BORDER
        ws.cell(row=row, column=stop_column).alignment = STOP_ALIGNMENT

        ws.cell(row=row, column=black_column).value = BLACK_HEADING
        ws.cell(row=row, column=black_column).border = HEADER_BORDER
        ws.cell(row=row, column=black_column).alignment = CENTER_ALIGNMENT

        ws.cell(row=row, column=white_column).value = WHITE_HEADING
        ws.cell(row=row, column=white_column).border = HEADER_BORDER
        ws.cell(row=row, column=white_column).alignment = CENTER_ALIGNMENT

        ws.cell(row=row, column=clue_column).value = CLUE_HEADING
        ws.cell(row=row, column=clue_column).border = HEADER_BORDER
        ws.cell(row=row, column=clue_column).alignment = CENTER_ALIGNMENT

        row += 1
        for line in range(self.args.stops):
            ws.cell(row=row + line,
                    column=stop_column).value = line + 1
            ws.cell(row=row + line,
                    column=stop_column).border = CELL_BORDER
            ws.cell(row=row + line,
                    column=stop_column).alignment = STOP_ALIGNMENT

            ws.cell(row=row + line,
                    column=black_column).border = CELL_BORDER
            ws.cell(row=row + line,
                    column=white_column).border = CELL_BORDER

            code = replacement.generate_clue(line + 1, self.clue_answers[line])
            ws.cell(row=row + line, column=clue_column).value = code
            ws.cell(row=row + line, column=clue_column).border = CELL_BORDER

            for column in range(self.args.columns):
                cell = ws.cell(row=row + line, column=2 + column)
                cell.value = self.clue_lines[line][column]
                cell.border = CELL_BORDER
                cell.alignment = COLOR_ALIGNMENT
                cell.fill = PatternFill(
                    "solid",
                    fgColor=Color(indexed=8 + self.clue_lines[line][column]))

        row += line

        assert row - start_row < ROWS_PER_SHEET


class Stops(object):
    """Maintains all informations from all stops as gotten from each sheet.

    Generates clues as information is added acting as the replacement
    object when creating sheets."""

    def __init__(self, args):
        self.args = args
        self.stop_infos = dict()
        self.max_clues = self.args.stops * self.args.sheets
        self.next_clue = self.generate_clues()

    def generate_clues(self):
        clues = list(range(100, 100 + self.max_clues))
        random.shuffle(clues)
        for c in clues:
            yield c

    def generate_clue(self, stop, tuple):
        if stop not in self.stop_infos:
            self.stop_infos[stop] = dict()
        if tuple not in self.stop_infos[stop]:
            self.stop_infos[stop][tuple] = next(self.next_clue)
        return self.stop_infos[stop][tuple]

    def output(self, ws, start_row):
        for stop_number, _ in enumerate(range(self.args.stops), start=1):
            ws.merge_cells(start_row=start_row, end_row=start_row,
                           start_column=1, end_column=9)
            ws.cell(row=start_row,
                    column=1).value = HEADING_PER_STOP + " " + str(stop_number)
            ws.cell(row=start_row,
                    column=1).alignment = INTRO_ALIGNMENT
            row = start_row + 3

            intro_lines = 6
            ws.merge_cells(start_row=row, end_row=row + intro_lines,
                           start_column=1, end_column=9)
            ws.cell(row=row, column=1).value = INTRO_TEXT_PER_STOP
            ws.cell(row=row, column=1).alignment = INTRO_ALIGNMENT

            row += intro_lines + 2
            clue_column = 1
            black_column = 2
            white_column = 3

            ws.cell(row=row, column=clue_column).value = CLUE_HEADING
            ws.cell(row=row, column=clue_column).border = HEADER_BORDER
            ws.cell(row=row, column=clue_column).alignment = CLUE_ALIGNMENT

            ws.cell(row=row, column=black_column).value = BLACK_HEADING
            ws.cell(row=row, column=black_column).border = HEADER_BORDER
            ws.cell(row=row, column=black_column).alignment = CLUE_ALIGNMENT

            ws.cell(row=row, column=white_column).value = WHITE_HEADING
            ws.cell(row=row, column=white_column).border = HEADER_BORDER
            ws.cell(row=row, column=white_column).alignment = CLUE_ALIGNMENT

            row += 2
            for clue, tuple in sorted([(v, k,)
                                       for k, v
                                       in self.stop_infos[stop_number].items()
                                       ]):
                blacks, whites = tuple
                ws.cell(row=row,
                        column=clue_column).value = str(clue)
                ws.cell(row=row,
                        column=clue_column).border = CELL_BORDER
                ws.cell(row=row,
                        column=clue_column).alignment = CLUE_ALIGNMENT

                ws.cell(row=row,
                        column=black_column).value = str(blacks)
                ws.cell(row=row,
                        column=black_column).border = CELL_BORDER
                ws.cell(row=row,
                        column=black_column).alignment = CLUE_ALIGNMENT

                ws.cell(row=row,
                        column=white_column).value = str(whites)
                ws.cell(row=row,
                        column=white_column).border = CELL_BORDER
                ws.cell(row=row,
                        column=white_column).alignment = CLUE_ALIGNMENT

                row += 1

            assert row - start_row < ROWS_PER_SHEET
            start_row += ROWS_PER_SHEET


if __name__ == "__main__":
    args = parser.parse_args()

    wb = openpyxl.Workbook()
    ws = wb.active

    stops = Stops(args)

    row = 1
    correct_lines = dict()
    for index in range(args.sheets):
        sheet_number = 1 + index
        s = Sheet(args, index < args.easy)
        print(s.correct)
        print(s.clue_lines)
        correct_lines[sheet_number] = (s.correct, s.solvable)

        s.output(ws, str(sheet_number), row, stops)
        row += ROWS_PER_SHEET

    correct_answers_heading_written = False
    line = 0
    for sheet_number, tuple in correct_lines.items():
        correct, solvable = tuple
        if not correct_answers_heading_written:
            correct_answers_heading_written = True
            ws.cell(row=row,
                    column=1).value = HEADING_CORRECT_ANSWERS
            ws.cell(row=row + 2,
                    column=1).value = CORRECT_HEADING
            ws.cell(row=row + 2,
                    column=2 + args.columns + 1).value = SOLVED_IN_HEADING
            line = 3

        ws.cell(row=row + line, column=1).value = sheet_number
        for column in range(args.columns):
            cell = ws.cell(row=row + line, column=2 + column)
            cell.value = correct[column]
            cell.border = CELL_BORDER
            cell.alignment = COLOR_ALIGNMENT
            cell.fill = PatternFill("solid",
                                    fgColor=Color(indexed=8 + correct[column]))
        ws.cell(row=row + line, column=2 + args.columns + 1).value = solvable

        line += 1
        if line > ROWS_PER_SHEET - 5:
            correct_answers_heading_written = False
            row += ROWS_PER_SHEET

    row += ROWS_PER_SHEET

    stops.output(ws, row)

    wb.save("mm.xlsx")
