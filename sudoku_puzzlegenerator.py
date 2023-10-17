#!/usr/bin/env python3

"""Generates a set of sudokus to solve i.e. an Excel-workbook to print out
with both participants' sheets and the stops.

Easy:
--clues-left=30

Hard:
--clues-left=0 --clue-letters=BCDEGPTV
"""

import argparse
import openpyxl
import openpyxl.styles
import random


HEADING_PER_SHEET = "Deltagarblankett"

INTRO_TEXT_PER_SHEET = """\
Det här är ett deltagarprotokoll för Radiosudoku på Skogsrå.

Regler för vanlig sudoku gäller dvs, i varje ruta ska det in en siffra
på ett sånt sätt att varje rad, varje kolumn, och varje 3x3-ruta
innehåller siffrorna 1 - 9.  Siffrorna för de rutor som har ledtrådar
längst upp istället hittar ni på någon av kontrollerna som är gömda runt
lägerområdena.  Ledtrådens siffra är vilken kontroll det är och
bokstavskoden är vilken ledtråd det är på den kontrollen."""

HEADING_PER_STOP = "Radiosudokukontroll nummer"

INTRO_TEXT_PER_STOP = """Detta är en kontroll för Radiosudoku på Skogsrå.

Det är en patrullövning med radiokommunikation som ni kan få göra som
en programaktivitet om ni bokar ett aktivitetspass med radioscouting
eller som ni kan prova en kväll."""

ROWS_PER_SHEET = 51


parser = argparse.ArgumentParser(description="Generate a set of sudoku games.")
parser.add_argument('--sheets', type=int, default=2,
                    help='Number of sheets')
parser.add_argument('--initial-values', type=int, default=15,
                    help='Number of values on each sheet initially')
parser.add_argument('--stops', type=int, default=13,
                    help='Number of stops to go to on the course')
parser.add_argument('--clue-letters', type=str,
                    default="ABCDEFGHIJKLMNOPQRSTUVWXYZ",
                    help='The letters to choose the clues from. '
                    'Must all be different')
parser.add_argument('--filename', type=str,
                    help='The file where the result is stored',
                    default="sudoku.xlsx")
parser.add_argument('--debug', '-d', action='store_true',
                    help='Activate trace outputs',
                    default=False)

args = None

CELL_ALIGNMENT = openpyxl.styles.Alignment(horizontal="center")
CLUE_ALIGNMENT = openpyxl.styles.Alignment(horizontal="right")
INTRO_ALIGNMENT = openpyxl.styles.Alignment(vertical="top",
                                            wrap_text=True)

CLUE_FONT = openpyxl.styles.Font(size=9, italic=True)
VALUE_FONT = openpyxl.styles.Font(size=14, bold=True)
CELL_SIDE = openpyxl.styles.Side(border_style="thin",
                                 color='FF000000')
BOX_SIDE = openpyxl.styles.Side(border_style="double",
                                color='FF000000')


class Replacement(object):
    def __init__(self, replacements, stops_for_clue):
        self._replacements = replacements
        self._stops_for_clue = stops_for_clue

    def get_clue(self, entry):
        return self._replacements.get(entry, None)

    def get_stop(self, entry):
        clue = self.get_clue(entry)
        if clue:
            return self._stops_for_clue.get(clue, None)
        return None

    def move_stops(self, translation):
        """Generate a new replacement.

        In the new replacement, the stops have new names according
        to translation.
        """
        return Replacement(self._replacements,
                           {key: translation[value]
                            for key, value
                            in self._stops_for_clue.items()})


class Sheet(object):
    def _found_in_row(self, candidate, index):
        """If the same number is already in the row, to the left,
        return True"""
        for i in range(9 * (index // 9), index):
            if candidate == self._board[i]:
                return True
        return False

    def _found_in_column(self, candidate, index):
        """If the same number is in the column above, return True"""
        column = index % 9
        for line in range(0, index // 9):
            if candidate == self._board[line * 9 + column]:
                return True
        return False

    def _found_in_box(self, candidate, index):
        """If the same number is already in the same box,
        above the index, return True"""
        box_column = (index % 9) // 3
        for line in range(3 * ((index // 9) // 3), index // 9):
            for col in range(box_column * 3, box_column * 3 + 3):
                if candidate == self._board[line * 9 + col]:
                    return True
        return False

    def _fill_board(self, index):
        if index >= 9 * 9:
            return True
        candidates = [c for c in range(1, 10)]
        random.shuffle(candidates)
        # import pdb; pdb.set_trace()
        for n in candidates:
            if self._found_in_row(n, index):
                continue
            if self._found_in_column(n, index):
                continue
            if self._found_in_box(n, index):
                continue
            if len(self._board) <= index:
                self._board.append(n)
            else:
                self._board[index] = n
            if self._fill_board(index + 1):
                return True
        return False

    def __init__(self):
        """Creates a fully filled sheet."""
        self._board = []
        self._fill_board(0)
        self.fully_filled_board = self._board.copy()

    def __eq__(self, other):
        return self.fully_filled_board == other.fully_filled_board

    def empty_random_cell(self):
        pos = random.randint(0, 9 * 9 - 1)
        if self._board[pos] == 0:
            self.empty_random_cell()
        self._board[pos] = 0

    def replace_by(self, clue):
        pos = random.randint(0, 9 * 9 - 1)
        if self._board[pos] not in range(1, 10):
            return self.replace_by(clue)
        value = self._board[pos]
        self._board[pos] = clue
        return value

    def get_board_value(self, line, column):
        entry = self._board[line * 9 + column]
        if entry == 0:
            return ""
        if entry > 9:
            return ""
        return entry

    def get_board_clue(self, line, column, replacement):
        entry = self._board[line * 9 + column]
        if entry == 0:
            return ""
        if entry <= 9:
            return ""
        clue = replacement.get_clue(entry)
        stop = replacement.get_stop(entry)
        if stop:
            return str(stop) + " " + clue
        return clue

    def print(self, replacement=None):
        for line in range(9):
            for column in range(9):
                entry = self._board[line * 9 + column]
                if entry == 0:
                    entry = ""
                elif replacement:
                    clue = replacement.get_clue(entry)
                    if clue:
                        stop = replacement.get_stop(entry)
                        if stop:
                            entry = str(stop) + " " + clue
                        else:
                            entry = clue
                print(f"{entry:>6}", end=' ')
            print()
        print()

    def print2(self, replacement):
        for line in range(9):
            for column in range(9):
                print(f"| {self.get_board_clue(line, column, replacement):^5}",
                      '/',
                      f"{self.get_board_value(line, column):^1}",
                      end=' ')
            print("|")
        print()

    def output(self, ws, sheet_identity, start_row, replacement):
        """Will fill the worksheet from line start_line with the sheet.

        After a common header with explanation, there board is generated
        with each cell as three spreadsheet cells on top of each-other:
        --------------
        |    clue    |
        --------------
        |   value    |
        --------------
        |            |
        --------------
        Either clue or value is given. No lines within each cell.
        Thin lines around except for the box lines that are double size.
        """
        ws.merge_cells(start_row=start_row, end_row=start_row,
                       start_column=1, end_column=9)
        ws.cell(row=start_row, column=1).value = sheet_identity
        ws.cell(row=start_row, column=1).alignment = INTRO_ALIGNMENT

        row = start_row + 3
        ws.merge_cells(start_row=row, end_row=row + 8,
                       start_column=1, end_column=9)
        ws.cell(row=row, column=1).value = INTRO_TEXT_PER_SHEET
        ws.cell(row=row, column=1).alignment = INTRO_ALIGNMENT

        row = start_row + 15
        for line in range(9):
            line_borders = [{"top": CELL_SIDE}, {}, {"bottom": CELL_SIDE}]
            if line % 3 == 0:
                line_borders[0]["top"] = BOX_SIDE
            if line % 3 == 2:
                line_borders[2]["bottom"] = BOX_SIDE
            for column in range(9):
                column_borders = {"left": CELL_SIDE, "right": CELL_SIDE}
                if column % 3 == 0:
                    column_borders["left"] = BOX_SIDE
                if column % 3 == 2:
                    column_borders["right"] = BOX_SIDE
                cell = ws.cell(row=row, column=1 + column)
                cell.value = self.get_board_clue(line, column, replacement)
                cell.font = CLUE_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = openpyxl.styles.Border(**line_borders[0],
                                                     **column_borders)

                cell = ws.cell(row=row + 1, column=1 + column)
                cell.value = self.get_board_value(line, column)
                cell.font = VALUE_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = openpyxl.styles.Border(**line_borders[1],
                                                     **column_borders)

                cell = ws.cell(row=row + 2, column=1 + column)
                cell.border = openpyxl.styles.Border(**line_borders[2],
                                                     **column_borders)

            row = row + 3

        assert row - start_row < ROWS_PER_SHEET


class SudokuGenerator(object):
    """Generator for the set of sheets and stops."""

    EMPTIED_CELLS = 10

    def __init__(self, number_of_sheets, initial_values,
                 number_of_stops,
                 clue_letters):
        self._number_of_sheets = number_of_sheets
        self._number_of_clues_per_sheet = (9 * 9
                                           - self.EMPTIED_CELLS
                                           - initial_values)
        self._number_of_stops = number_of_stops
        self._clue_letters = clue_letters

        # Known weakness: If the clue_letters contains repeats
        # then this will not be enough and clue_generator will
        # search indefinately for options
        for i in range(1, 10):
            if (pow(len(self._clue_letters), i) >
                    self._number_of_sheets * self._number_of_clues_per_sheet):
                self._clue_length = i
                break
        else:
            print("Too many clues")
        self._used_clues = []
        self.sheets = []

        self._replacements = dict()        # value => [entry, ...]
        self._replacement_clues = dict()   # entry => clue
        self.stops = dict()                # stop# => [(clue, value), ...]
        self._stop_for_clue = dict()       # clue => stopindex

    def generate_clue(self):
        clue = ""
        for i in range(self._clue_length):
            r = random.randint(0, len(self._clue_letters) - 1)
            clue = clue + self._clue_letters[r]
        if clue in self._used_clues:
            return self.generate_clue()
        self._used_clues.append(clue)
        return clue

    def allocate_replacements_to_stops(self):
        """Create a list of stops in self.stops.
        Each stop is a list of clue => value pairs.
        Clues to replace the stop with is self._replacement_clues.
        On what stop each clue is is in self._stop_for_clue.
        """

        global args
        heaps_per_stop = 7 ** (1/2)
        saved_heaps = dict()
        for entry, replacements in self._replacements.items():
            random.shuffle(replacements)
            heaps = []
            for i in range(int(self._number_of_stops * heaps_per_stop)):
                heaps.append([])
            for replacement in replacements:
                heaps[random.randint(0, len(heaps) - 1)].append(replacement)
            heaps = sorted([x for x in heaps if x], key=len, reverse=True)
            saved_heaps[entry] = heaps

        if args.debug:
            print("Saved heaps:", saved_heaps)
            for key, value in saved_heaps.items():
                print(key, len(value), end=': ')
                for heap in value:
                    print(len(heap), end=' ')
                    print()

        saved_heaps_tuples = []
        for key, value in saved_heaps.items():
            for heap in value:
                clue = self.generate_clue()
                saved_heaps_tuples.append((heap, key, clue))
                for entry in heap:
                    self._replacement_clues[entry] = clue

        # The value with the heap with most entries is first.
        saved_heaps_tuples = sorted(saved_heaps_tuples,
                                    key=lambda x: len(x[0]),
                                    reverse=True)
        if args.debug:
            print("Saved heaps' tuples:", saved_heaps_tuples)

        stops_tuples = [[] for _ in range(self._number_of_stops)]
        while saved_heaps_tuples:
            # allocated it to the stop with the least entries
            stops_tuples = sorted(stops_tuples,
                                  key=lambda x: sum([len(t[0]) for t in x]))
            stops_tuples[0].append(saved_heaps_tuples.pop(0))

        if args.debug:
            print("Stops' tuples:", stops_tuples)
            for tuples in stops_tuples:
                print("Heaps:", len(tuples),
                      "Values:", sum([len(x[0]) for x in tuples]))

        # allocate one clue per heap

        self.stops = []
        for stop in stops_tuples:
            self.stops.append(sorted([(clue, value)
                                      for heap, value, clue in stop]))

        random.shuffle(self.stops)
        for n, stop in enumerate(self.stops):
            for clue, _ in stop:
                self._stop_for_clue[clue] = n

    def calculate(self):
        """Generate a set of sheets then move clues to stops."""
        clue = 100
        for i in range(self._number_of_sheets):

            sheet = Sheet()
            if sheet in self.sheets:
                print("That soduko is already seen")
                # If this happens, it means that we will get fewer
                # sheets than we actally wanted
                continue
            self.sheets.append(sheet)
            for c in range(self.EMPTIED_CELLS):
                sheet.empty_random_cell()
            for c in range(self._number_of_clues_per_sheet):
                clue = clue + 1
                value = sheet.replace_by(clue)
                if value not in self._replacements:
                    self._replacements[value] = []
                self._replacements[value].append(clue)

        self.allocate_replacements_to_stops()
        self.replacement = Replacement(self._replacement_clues,
                                       self._stop_for_clue)

    def output_stop(self, ws, stop, stop_number, stop_identity, start_row):
        ws.merge_cells(start_row=start_row, end_row=start_row,
                       start_column=1, end_column=9)
        ws.cell(row=start_row, column=1).value = stop_identity
        ws.cell(row=start_row, column=1).alignment = INTRO_ALIGNMENT

        row = start_row + 3
        ws.merge_cells(start_row=row, end_row=row + 6,
                       start_column=1, end_column=9)
        ws.cell(row=row, column=1).value = INTRO_TEXT_PER_STOP
        ws.cell(row=row, column=1).alignment = INTRO_ALIGNMENT

        start_row = row + 8
        CLUES_PER_COLUMN = 15
        for n, tuple in enumerate(stop):
            clue, value = tuple
            row = start_row + (n % CLUES_PER_COLUMN) * 2
            column = 1 + 4 * (n // CLUES_PER_COLUMN)
            ws.cell(row=row, column=column).value = clue + " :"
            ws.cell(row=row, column=column).alignment = CLUE_ALIGNMENT

            ws.cell(row=row, column=column + 1).value = value
            ws.cell(row=row, column=column + 1).alignment = CELL_ALIGNMENT


if __name__ == "__main__":
    args = parser.parse_args()

    gen = SudokuGenerator(args.sheets, args.initial_values,
                          args.stops,
                          args.clue_letters)

    gen.calculate()
    print('Sheets:')
    for s in gen.sheets:
        s.print(gen.replacement)
    print('Stops:')
    for n, s in enumerate(gen.stops):
        print(n, ", ".join([c+":"+str(con) for c, con in s]))
    print()

    wb = openpyxl.Workbook()
    ws = wb.active

    # Number the stops from 1 instead of from 0
    stop_number_translation = {x: y
                               for x, y
                               in enumerate(range(1, 1 + args.stops))}
    replacement = gen.replacement.move_stops(stop_number_translation)

    start_row = 1  # In spreadsheet indexing this is the first
    for n, s in enumerate(gen.sheets):
        s.output(ws, HEADING_PER_SHEET + " " + str(1 + n),
                 start_row, replacement)
        start_row = start_row + ROWS_PER_SHEET

    for n, s in enumerate(gen.stops):
        stop_number = stop_number_translation[n]
        gen.output_stop(ws, s, stop_number,
                        HEADING_PER_STOP + " " + str(stop_number),
                        start_row)
        start_row = start_row + ROWS_PER_SHEET

    wb.save(args.filename)

    print('Saved output in', args.filename)
    print('To print,')
    print('1. open in Excel or LibreOffice Calc,')
    print('2. adjust page size so that all 9 columns of each sudoku is seen')
    print('   and so that the pages has the right height (the first line on')
    print('   every page is in the same place)')
    print('3. Print (or export to pdf)')
